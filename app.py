
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from datetime import datetime
import io
import csv

from openpyxl import load_workbook, Workbook

from database import get_connection, init_db

app = Flask(__name__)
app.secret_key = "redutron-secret-key"

init_db()


# ---------- helpers ----------

def parse_float_br(value):
    if value is None:
        return 0.0
    v = str(value).strip()
    if not v or v.lower() == "nan":
        return 0.0
    v = v.replace("R$", "").replace(" ", "")
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    else:
        v = v.replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    v = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    return v


def get_percentuais():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT key, value FROM settings")
    rows = cur.fetchall()
    conn.close()
    data = {r["key"]: r["value"] for r in rows}
    try:
        imposto = float(data.get("imposto_pct", "0.05"))
    except ValueError:
        imposto = 0.05
    try:
        despesa = float(data.get("despesa_pct", "0.035"))
    except ValueError:
        despesa = 0.035
    return imposto, despesa


def get_inventory(product_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT quantity, avg_cost FROM product_inventory WHERE product_id = ?", (product_id,))
    row = cur.fetchone()
    if row is None:
        cur.execute(
            "INSERT INTO product_inventory (product_id, quantity, avg_cost) VALUES (?, 0, 0)",
            (product_id,),
        )
        conn.commit()
        quantity, avg_cost = 0.0, 0.0
    else:
        quantity, avg_cost = row["quantity"], row["avg_cost"]
    conn.close()
    return quantity, avg_cost


def update_inventory_entry(product_id, quantity, cost_unit, date, origin="entrada"):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO stock_entries (product_id, quantity, cost_unit, date, origin)
        VALUES (?, ?, ?, ?, ?)
        """,
        (product_id, quantity, cost_unit, date, origin),
    )
    cur.execute(
        "SELECT quantity, avg_cost FROM product_inventory WHERE product_id = ?",
        (product_id,),
    )
    row = cur.fetchone()
    if row is None:
        old_q, old_cost = 0.0, 0.0
        cur.execute(
            "INSERT INTO product_inventory (product_id, quantity, avg_cost) VALUES (?, 0, 0)",
            (product_id,),
        )
    else:
        old_q, old_cost = row["quantity"], row["avg_cost"]
    new_q = old_q + quantity
    if new_q > 0:
        new_cost = (old_q * old_cost + quantity * cost_unit) / new_q
    else:
        new_cost = 0.0
    cur.execute(
        "UPDATE product_inventory SET quantity = ?, avg_cost = ? WHERE product_id = ?",
        (new_q, new_cost, product_id),
    )
    conn.commit()
    conn.close()


def update_inventory_sale(product_id, quantity):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT quantity FROM product_inventory WHERE product_id = ?", (product_id,))
    row = cur.fetchone()
    if row is None:
        cur.execute(
            "INSERT INTO product_inventory (product_id, quantity, avg_cost) VALUES (?, ?, 0)",
            (product_id, -quantity),
        )
    else:
        new_q = (row["quantity"] or 0) - quantity
        cur.execute(
            "UPDATE product_inventory SET quantity = ? WHERE product_id = ?",
            (new_q, product_id),
        )
    conn.commit()
    conn.close()


# ---------- rotas básicas ----------

@app.route("/")
def index():
    return render_template("index.html")


# ---------- produtos ----------

@app.route("/produtos")
def produtos():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products ORDER BY name")
    products = cur.fetchall()
    conn.close()
    return render_template("products.html", products=products)


@app.route("/produtos/novo", methods=["POST"])
def novo_produto():
    name = request.form.get("name", "").strip()
    sku = request.form.get("sku") or None
    variable_cost = parse_float_br(request.form.get("variable_cost"))
    default_price = parse_float_br(request.form.get("default_price"))
    if not name:
        flash("Nome é obrigatório.", "error")
        return redirect(url_for("produtos"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO products (name, sku, variable_cost, default_price) VALUES (?, ?, ?, ?)",
        (name, sku, variable_cost, default_price),
    )
    conn.commit()
    conn.close()
    flash("Produto cadastrado.", "success")
    return redirect(url_for("produtos"))


@app.route("/produtos/<int:product_id>/editar", methods=["GET", "POST"])
def editar_produto(product_id):
    conn = get_connection()
    cur = conn.cursor()
    if request.method == "GET":
        cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
        p = cur.fetchone()
        conn.close()
        if not p:
            flash("Produto não encontrado.", "error")
            return redirect(url_for("produtos"))
        return render_template("product_edit.html", p=p)
    name = request.form.get("name", "").strip()
    sku = request.form.get("sku") or None
    variable_cost = parse_float_br(request.form.get("variable_cost"))
    default_price = parse_float_br(request.form.get("default_price"))
    cur.execute(
        "UPDATE products SET name = ?, sku = ?, variable_cost = ?, default_price = ? WHERE id = ?",
        (name, sku, variable_cost, default_price, product_id),
    )
    conn.commit()
    conn.close()
    flash("Produto atualizado.", "success")
    return redirect(url_for("produtos"))


@app.route("/produtos/<int:product_id>/deletar", methods=["POST"])
def deletar_produto(product_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM sales WHERE product_id = ?", (product_id,))
    cur.execute("DELETE FROM stock_entries WHERE product_id = ?", (product_id,))
    cur.execute("DELETE FROM product_inventory WHERE product_id = ?", (product_id,))
    cur.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()
    conn.close()
    flash("Produto e dados relacionados deletados.", "success")
    return redirect(url_for("produtos"))


# ---------- vendas manuais ----------

@app.route("/vendas")
def vendas():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT s.*, p.name AS product_name
        FROM sales s
        JOIN products p ON p.id = s.product_id
        ORDER BY s.date DESC, s.id DESC
        """
    )
    sales = cur.fetchall()
    cur.execute("SELECT id, name FROM products ORDER BY name")
    products = cur.fetchall()
    conn.close()
    return render_template("sales.html", sales=sales, products=products)


@app.route("/vendas/novo", methods=["POST"])
def nova_venda():
    try:
        product_id = int(request.form["product_id"])
        date = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
        quantity = parse_float_br(request.form.get("quantity"))
        unit_price = parse_float_br(request.form.get("unit_price"))
        marketplace_fee = parse_float_br(request.form.get("marketplace_fee"))
        other_variable_cost = parse_float_br(request.form.get("other_variable_cost"))
        discount = parse_float_br(request.form.get("discount"))
    except Exception:
        flash("Erro ao ler dados.", "error")
        return redirect(url_for("vendas"))
    if quantity <= 0:
        flash("Quantidade deve ser > 0.", "error")
        return redirect(url_for("vendas"))
    _, avg_cost = get_inventory(product_id)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO sales (
            product_id, date, quantity, unit_price,
            marketplace_fee, other_variable_cost, discount,
            cost_unit_at_sale, source
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            product_id,
            parse_date(date),
            quantity,
            unit_price,
            marketplace_fee,
            other_variable_cost,
            discount,
            avg_cost,
            "manual",
        ),
    )
    conn.commit()
    conn.close()
    update_inventory_sale(product_id, quantity)
    flash("Venda lançada.", "success")
    return redirect(url_for("vendas"))


# ---------- estoque ----------

@app.route("/estoque")
def estoque():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            p.id,
            p.name,
            p.sku,
            IFNULL(pi.quantity, 0) AS quantity,
            IFNULL(pi.avg_cost, 0) AS avg_cost
        FROM products p
        LEFT JOIN product_inventory pi ON pi.product_id = p.id
        ORDER BY p.name
        """
    )
    rows = cur.fetchall()
    conn.close()
    dados = []
    for r in rows:
        total = (r["quantity"] or 0) * (r["avg_cost"] or 0)
        dados.append(
            {
                "id": r["id"],
                "name": r["name"],
                "sku": r["sku"],
                "quantity": r["quantity"],
                "avg_cost": r["avg_cost"],
                "total": total,
            }
        )
    return render_template("estoque.html", dados=dados)


@app.route("/estoque/entrada", methods=["GET", "POST"])
def entrada_estoque():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM products ORDER BY name")
    products = cur.fetchall()
    conn.close()
    if request.method == "GET":
        return render_template("entrada_estoque.html", products=products)
    try:
        product_id = int(request.form["product_id"])
        date = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
        quantity = parse_float_br(request.form.get("quantity"))
        cost_unit = parse_float_br(request.form.get("cost_unit"))
        origin = request.form.get("origin") or "entrada_manual"
    except Exception:
        flash("Erro ao ler dados.", "error")
        return redirect(url_for("entrada_estoque"))
    if quantity <= 0 or cost_unit <= 0:
        flash("Qtd e custo precisam ser > 0.", "error")
        return redirect(url_for("entrada_estoque"))
    update_inventory_entry(product_id, quantity, cost_unit, parse_date(date), origin)
    flash("Entrada registrada.", "success")
    return redirect(url_for("estoque"))


@app.route("/estoque/ajuste", methods=["GET", "POST"])
def ajuste_estoque():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM products ORDER BY name")
    products = cur.fetchall()
    conn.close()
    if request.method == "GET":
        return render_template("ajuste_estoque.html", products=products)
    try:
        product_id = int(request.form["product_id"])
        new_quantity = parse_float_br(request.form.get("new_quantity"))
    except Exception:
        flash("Erro ao ler dados.", "error")
        return redirect(url_for("ajuste_estoque"))
    current_q, avg_cost = get_inventory(product_id)
    diff = new_quantity - current_q
    if diff == 0:
        flash("Nenhuma alteração necessária.", "info")
        return redirect(url_for("estoque"))
    today = datetime.now().strftime("%Y-%m-%d")
    if diff > 0:
        update_inventory_entry(product_id, diff, avg_cost or 0, today, origin="ajuste_positivo")
    else:
        update_inventory_sale(product_id, -diff)
    flash("Ajuste de estoque realizado.", "success")
    return redirect(url_for("estoque"))


# ---------- configurações ----------

@app.route("/configuracoes", methods=["GET", "POST"])
def configuracoes():
    if request.method == "POST":
        imposto_pct = parse_float_br(request.form.get("imposto_pct")) / 100.0
        despesa_pct = parse_float_br(request.form.get("despesa_pct")) / 100.0
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO settings (key, value) VALUES ('imposto_pct', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (str(imposto_pct),),
        )
        cur.execute(
            "INSERT INTO settings (key, value) VALUES ('despesa_pct', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (str(despesa_pct),),
        )
        conn.commit()
        conn.close()
        flash("Configurações salvas.", "success")
        return redirect(url_for("configuracoes"))
    imposto, despesa = get_percentuais()
    return render_template(
        "configuracoes.html",
        imposto_pct=imposto * 100,
        despesa_pct=despesa * 100,
    )


# ---------- relatórios ----------

def calcular_relatorio(inicio, fim, criterio="faturamento"):
    imposto_pct, despesa_pct = get_percentuais()
    conn = get_connection()
    cur = conn.cursor()
    where = "WHERE 1=1"
    params = []
    if inicio:
        where += " AND s.date >= ?"
        params.append(inicio)
    if fim:
        where += " AND s.date <= ?"
        params.append(fim)
    cur.execute(
        f"""
        SELECT
            p.id AS product_id,
            p.name AS product_name,
            SUM(s.quantity) AS total_qty,
            SUM(s.quantity * s.unit_price) AS faturamento_bruto,
            SUM(s.discount) AS total_desconto,
            SUM(s.marketplace_fee) AS total_marketplace_fee,
            SUM(s.other_variable_cost) AS total_other_cost,
            SUM(s.quantity * s.cost_unit_at_sale) AS total_custo_produto
        FROM sales s
        JOIN products p ON p.id = s.product_id
        {where}
        GROUP BY p.id, p.name
        """
        ,
        params,
    )
    rows = cur.fetchall()
    conn.close()
    rel = []
    total_bruto = 0.0
    totals = {
        "qtd": 0.0,
        "desconto": 0.0,
        "marketplace": 0.0,
        "outros": 0.0,
        "imposto": 0.0,
        "despesa": 0.0,
        "custo_produto": 0.0,
        "custos_var": 0.0,
        "margem": 0.0,
    }
    for r in rows:
        bruto = r["faturamento_bruto"] or 0.0
        desconto = r["total_desconto"] or 0.0
        comissao = r["total_marketplace_fee"] or 0.0
        outros = r["total_other_cost"] or 0.0
        custo_prod = r["total_custo_produto"] or 0.0
        imposto = bruto * imposto_pct
        base_desp = bruto - comissao
        despesa = base_desp * despesa_pct
        liquido = bruto - desconto
        custos_var = custo_prod + comissao + outros + imposto + despesa
        margem = liquido - custos_var
        margem_pct = (margem / liquido * 100) if liquido > 0 else 0.0
        total_bruto += bruto
        totals["qtd"] += r["total_qty"] or 0
        totals["desconto"] += desconto
        totals["marketplace"] += comissao
        totals["outros"] += outros
        totals["imposto"] += imposto
        totals["despesa"] += despesa
        totals["custo_produto"] += custo_prod
        totals["custos_var"] += custos_var
        totals["margem"] += margem
        rel.append(
            {
                "product_name": r["product_name"],
                "total_qty": r["total_qty"] or 0,
                "faturamento_bruto": bruto,
                "total_desconto": desconto,
                "marketplace_fee": comissao,
                "other_cost": outros,
                "imposto": imposto,
                "despesa": despesa,
                "custo_produto": custo_prod,
                "total_custos_variaveis": custos_var,
                "margem_contrib": margem,
                "margem_pct": margem_pct,
            }
        )
    if total_bruto == 0:
        return [], 0.0, 0, totals
    if criterio == "margem":
        key_fn = lambda x: x["margem_contrib"]
    else:
        key_fn = lambda x: x["faturamento_bruto"]
    rel.sort(key=key_fn, reverse=True)
    cumul = 0.0
    for item in rel:
        cumul += item["faturamento_bruto"]
        perc_acum = cumul / total_bruto * 100
        if perc_acum <= 80:
            curva = "A"
        elif perc_acum <= 95:
            curva = "B"
        else:
            curva = "C"
        item["perc_acumulado"] = perc_acum
        item["curva"] = curva
    return rel, total_bruto, len(rel), totals


@app.route("/relatorio", methods=["GET", "POST"])
def relatorio():
    inicio = fim = None
    criterio = "faturamento"
    if request.method == "POST":
        inicio = request.form.get("inicio") or None
        fim = request.form.get("fim") or None
        criterio = request.form.get("criterio") or "faturamento"
    dados, total_bruto, total_itens, totals = calcular_relatorio(inicio, fim, criterio)
    return render_template(
        "report.html",
        dados=dados,
        total_fat_bruto=total_bruto,
        total_itens=total_itens,
        inicio=inicio,
        fim=fim,
        criterio=criterio,
        totals=totals,
    )


@app.route("/relatorio/csv")
def relatorio_csv():
    inicio = request.args.get("inicio") or None
    fim = request.args.get("fim") or None
    criterio = request.args.get("criterio") or "faturamento"
    dados, total_bruto, _, totals = calcular_relatorio(inicio, fim, criterio)
    output = io.StringIO()
    w = csv.writer(output, delimiter=";")
    w.writerow(
        [
            "Produto",
            "Quantidade",
            "Faturamento bruto",
            "Descontos",
            "Comissão",
            "Impostos",
            "Despesas",
            "Outros custos",
            "Custo produto",
            "Custos variáveis",
            "Margem contribuição",
            "Margem (%)",
            "Curva",
        ]
    )
    for d in dados:
        w.writerow(
            [
                d["product_name"],
                d["total_qty"],
                f"{d['faturamento_bruto']:.2f}",
                f"{d['total_desconto']:.2f}",
                f"{d['marketplace_fee']:.2f}",
                f"{d['imposto']:.2f}",
                f"{d['despesa']:.2f}",
                f"{d['other_cost']:.2f}",
                f"{d['custo_produto']:.2f}",
                f"{d['total_custos_variaveis']:.2f}",
                f"{d['margem_contrib']:.2f}",
                f"{d['margem_pct']:.2f}",
                d["curva"],
            ]
        )
    w.writerow([])
    w.writerow(["TOTAL QTD", f"{totals['qtd']:.2f}"])
    w.writerow(["TOTAL FATURAMENTO BRUTO", f"{total_bruto:.2f}"])
    w.writerow(["TOTAL DESCONTOS", f"{totals['desconto']:.2f}"])
    w.writerow(["TOTAL COMISSÃO", f"{totals['marketplace']:.2f}"])
    w.writerow(["TOTAL IMPOSTOS", f"{totals['imposto']:.2f}"])
    w.writerow(["TOTAL DESPESAS", f"{totals['despesa']:.2f}"])
    w.writerow(["TOTAL OUTROS CUSTOS", f"{totals['outros']:.2f}"])
    w.writerow(["TOTAL CUSTO PRODUTO", f"{totals['custo_produto']:.2f}"])
    w.writerow(["TOTAL CUSTOS VARIÁVEIS", f"{totals['custos_var']:.2f}"])
    w.writerow(["TOTAL MARGEM CONTRIBUIÇÃO", f"{totals['margem']:.2f}"])
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="relatorio_redutron.csv",
    )


# ---------- exportar modelo e importar vendas consolidadas ----------

@app.route("/vendas/modelo")
def exportar_modelo_vendas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas Consolidadas"
    ws.append(["Produto", "Data", "Quantidade", "Preço unitário", "Comissão", "Origem"])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="modelo_vendas_consolidadas.xlsx",
    )


@app.route("/vendas/importar_consolidado", methods=["GET", "POST"])
def importar_consolidado():
    if request.method == "GET":
        return render_template("import_consolidado.html")
    file = request.files.get("arquivo")
    if not file or file.filename == "":
        flash("Selecione um arquivo XLSX.", "error")
        return redirect(url_for("importar_consolidado"))
    from pandas import read_excel
    try:
        df = read_excel(file)
    except Exception:
        flash("Não foi possível ler o arquivo.", "error")
        return redirect(url_for("importar_consolidado"))
    required = ["Produto", "Data", "Quantidade", "Preço unitário", "Comissão"]
    for col in required:
        if col not in df.columns:
            flash(f"Coluna obrigatória ausente: {col}", "error")
            return redirect(url_for("importar_consolidado"))
    conn = get_connection()
    cur = conn.cursor()
    importados = 0
    for _, row in df.iterrows():
        produto_txt = str(row["Produto"]).strip()
        if not produto_txt or produto_txt.lower() in ("nan", "none"):
            continue
        data = parse_date(row["Data"])
        qtd = parse_float_br(row["Quantidade"])
        preco = parse_float_br(row["Preço unitário"])
        comissao = parse_float_br(row["Comissão"])
        origem = str(row.get("Origem") or "consolidado").strip()
        if qtd <= 0 or preco <= 0:
            continue
        cur.execute("SELECT id FROM products WHERE sku = ? OR name = ?", (produto_txt, produto_txt))
        p = cur.fetchone()
        if p:
            product_id = p["id"]
        else:
            cur.execute(
                "INSERT INTO products (name, sku, variable_cost, default_price) VALUES (?, ?, ?, ?)",
                (produto_txt, produto_txt, 0.0, preco),
            )
            product_id = cur.lastrowid
        _, avg_cost = get_inventory(product_id)
        cur.execute(
            """
            INSERT INTO sales (
                product_id, date, quantity, unit_price,
                marketplace_fee, other_variable_cost, discount,
                cost_unit_at_sale, source
            )
            VALUES (?, ?, ?, ?, ?, 0, 0, ?, ?)
            """,
            (
                product_id,
                data,
                qtd,
                preco,
                comissao,
                avg_cost,
                origem,
            ),
        )
        update_inventory_sale(product_id, qtd)
        importados += 1
    conn.commit()
    conn.close()
    flash(f"Importação consolidada concluída. Vendas importadas: {importados}.", "success")
    return redirect(url_for("vendas"))


# ---------- importação ML original ----------

@app.route("/importar", methods=["GET", "POST"])
def importar_ml():
    if request.method == "GET":
        return render_template("import.html")
    file = request.files.get("arquivo")
    if not file or file.filename == "":
        flash("Selecione um arquivo XLSX.", "error")
        return redirect(url_for("importar_ml"))
    data = file.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheet = wb.active
    header_row = 6
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val is None:
            continue
        headers[str(val).strip().lower()] = col

    def find_col(*names):
        for n in names:
            col = headers.get(n.lower())
            if col:
                return col
        return None

    col_sku = find_col("sku")
    col_data = find_col("data da venda")
    col_qtd = find_col("unidades", "quantidade", "quantidade / unidades")
    col_receita = find_col("receita por produtos (brl)", "receita por produtos")
    col_tarifa = find_col(
        "tarifa de venda e impostos (brl)",
        "tarifa de venda e impostos brl",
        "tarifa de venda e impostos",
    )
    col_anuncio = find_col("# de anúncio", "# de anuncio", "nº do anúncio", "n° do anúncio")
    col_titulo = find_col("título do anúncio", "titulo do anuncio")

    if not all([col_data, col_qtd, col_receita, col_tarifa]):
        flash("Não encontrei todas as colunas obrigatórias na planilha.", "error")
        return redirect(url_for("importar_ml"))

    conn = get_connection()
    cur = conn.cursor()
    importados = 0
    erros = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        try:
            data_v = sheet.cell(row=row_idx, column=col_data).value
            qtd_v = sheet.cell(row=row_idx, column=col_qtd).value
            rec_v = sheet.cell(row=row_idx, column=col_receita).value
            tarifa_v = sheet.cell(row=row_idx, column=col_tarifa).value
            if not qtd_v:
                continue
            qtd = parse_float_br(qtd_v)
            if qtd <= 0:
                continue
            data_str = parse_date(data_v)
            receita = parse_float_br(rec_v)
            tarifa = parse_float_br(tarifa_v)
            bruto = receita + tarifa
            unit_price = bruto / qtd if qtd > 0 else 0.0
            sku_v = sheet.cell(row=row_idx, column=col_sku).value if col_sku else None
            anuncio_v = sheet.cell(row=row_idx, column=col_anuncio).value if col_anuncio else None
            titulo_v = sheet.cell(row=row_idx, column=col_titulo).value if col_titulo else None
            sku = str(sku_v).strip() if sku_v else None
            if not sku and anuncio_v:
                sku = str(anuncio_v).strip()
            if not sku:
                sku = "SEM-SKU"
            cur.execute("SELECT id FROM products WHERE sku = ? OR name = ?", (sku, sku))
            p = cur.fetchone()
            if p:
                product_id = p["id"]
            else:
                nome_prod = titulo_v or f"Produto {sku}"
                cur.execute(
                    "INSERT INTO products (name, sku, variable_cost, default_price) VALUES (?, ?, ?, ?)",
                    (nome_prod, sku, 0.0, unit_price),
                )
                product_id = cur.lastrowid
            _, avg_cost = get_inventory(product_id)
            cur.execute(
                """
                INSERT INTO sales (
                    product_id, date, quantity, unit_price,
                    marketplace_fee, other_variable_cost, discount,
                    cost_unit_at_sale, source
                )
                VALUES (?, ?, ?, ?, ?, 0, 0, ?, ?)
                """,
                (
                    product_id,
                    data_str,
                    qtd,
                    unit_price,
                    tarifa,
                    avg_cost,
                    "ml",
                ),
            )
            update_inventory_sale(product_id, qtd)
            importados += 1
        except Exception:
            erros += 1
            continue
    conn.commit()
    conn.close()
    flash(
        f"Importação ML concluída. Vendas importadas: {importados}. Linhas com erro: {erros}.",
        "success" if importados else "error",
    )
    return redirect(url_for("vendas"))


if __name__ == "__main__":
    app.run(debug=True)
